# coding: utf-8

try:
  from openpyxl.workbook.workbook import Workbook
  from openpyxl.worksheet.worksheet import Worksheet
  import openpyxl
  from pathlib import Path
  import json
  from typing import List, Optional, Any, Dict, Tuple, Union
  import re
  import sys
  import os
  from calendar import monthrange

except (ImportError, ModuleNotFoundError):
  print("モジュールが読み込めません。フォルダの置き場所を間違えている可能性があります。")
  exit()
  
DIR_PATH = Path("エクセルリスト").resolve()
KLASS_NAME_FILE_PATH = Path("クラス名.txt").resolve()
STORAGE_DIR_PATH = Path("..\\ストレージ").resolve()

class AppError(Exception):
  def __init__(self, message: str):
    self.message = message

def debugPrint(explain, *args, **kwargs):
  strings = []
  for arg in args:
    strings.append(f"<{arg}>")
  for key, value in kwargs.items():
    strings.append(f"<{key}={value}>")
  print(f"### Debug ### {explain}: {' '.join(strings)} ### Debug ###", file=sys.stderr)

def abortProgram():
  print("エラーが発生したため、プログラムを中断します。", file=sys.stderr)
  exit()

def getNumberOfDays(reiwa_nendo: int, month: int) -> int:
  __AD_REIWA_GANNNENN = 2019
  reiwa_nen = reiwa_nendo if month >= 4 else reiwa_nendo + 1
  ad_nen = (__AD_REIWA_GANNNENN - 1) + reiwa_nen
  return monthrange(ad_nen, month)[1]

def checkPathsExistence():
  if not DIR_PATH.exists():
    raise AppError("エクセルリストフォルダが存在しません。")
  if not KLASS_NAME_FILE_PATH.exists():
    raise AppError("クラス名.txtが存在しません。")
  if not STORAGE_DIR_PATH.exists():
    raise AppError("ストレージフォルダが存在しません。")

def loadKlassName(klass_name_file_path: Path) -> str:
  try:
    with open(klass_name_file_path, "r", encoding="utf-8") as f:
      _klass = f.readline()
      klass = _klass.rstrip()
  except Exception as e:
    raise AppError("クラス名.txtの読み込みに失敗しました。")
  if klass == "" or klass.isspace():
    raise AppError("クラス名.txtにはクラス名を記述してください。")
  return klass

def readMonthlySave(storage_path: Path, reiwa: int, klass: str, month: int) -> List:
  reiwa_dir_path = storage_path.joinpath(str(reiwa)).resolve()
  if not reiwa_dir_path.exists():
    raise AppError(f"エラー: {reiwa}年のフォルダが存在しません。")
  klass_dir_path = reiwa_dir_path.joinpath(klass).resolve()
  if not klass_dir_path.exists():
    raise AppError(f"エラー: {reiwa}年{klass}組のフォルダが存在しません。")
  month_file_path = klass_dir_path.joinpath(str(month)).resolve()
  if not month_file_path.exists():
    raise AppError(f"エラー: {reiwa}年{klass}組{month}月のjsonファイルが存在しません。")
  
  with open(month_file_path, "r", encoding="utf-8") as f:
    return json.load(f)

class StructInspection:
  def __init__(self, name: str, description: str):
    self.name = name
    self.description = description
  
  def __repr__(self) -> str:
    return f"StructInspection({self.name}, {self.description})"

def splitInspections(raw: str) -> Tuple[List[StructInspection], List[str]]:
  warnings: List[str] = []
  splitted = raw.split(", ")
  inspection_list: List[StructInspection] = []
  for i in range(len(splitted)):
    element = splitted[i]
    #group by pattern matching and check if the element is in the format of `name: description`, then split it.
    pattern = r"(.+): (.+)"
    match = re.match(pattern, element, re.DOTALL)
    if match is None:
      warnings.append(f"注意(視診を処理中): {element}は正しい形式ではありません。 `名前: 説明`の形式である必要があります。 処理をスキップします。")
      continue
    name = match.group(1)
    description = match.group(2)
    inspection_list.append(StructInspection(name, description))
  return inspection_list, warnings
    

def splitActivities(raw: str) -> List[str]:
  splitted = raw.split(", ")
  return splitted

class StructDayFlow:
  def __init__(self, hour: int, minute: int, description: str):
    self.hour = hour
    self.minute = minute
    self.description = description
  
  def __repr__(self) -> str:
    return f"StructDayFlow({self.hour}, {self.minute}, {self.description})"

def splitDayFlows(raw: str) -> Tuple[List[StructDayFlow], List[str]]:
  warnings: List[str] = []
  splitted = raw.split(", ")
  # Now, as an example, `[9時30分]片付け`, 
  # each element is a string described in the format `[{hour}時{minute}分]{description}`, 
  # which is stored in StructDayFlow and returned as a list. 
  day_flow_list: List[StructDayFlow] = []
  for i in range(len(splitted)):
    element_str = splitted[i]
    element_str = element_str.lstrip()
    #group by pattern matching and check if the element is in the format of `[{hour}時{minute}分]{description}`, then extract hour, minute, description.
    pattern = r"\[(\d+)時(\d+)分\](.+)"
    match = re.match(pattern, element_str)
    if match is None:
      warnings.append(f"注意(活動の流れを処理中): {element_str}は正しい形式ではありません。 `[~時~分]<説明>`の形式である必要があります。 処理をスキップします。")
      continue
    hour_str = match.group(1)
    try:
      hour = int(hour_str)
    except ValueError as e:
      warnings.append(f"注意(活動の流れを処理中): {hour_str}は整数ではありません。 処理をスキップします。")
      continue
    minute_str = match.group(2)
    try:
      minute = int(minute_str)
    except ValueError as e:
      warnings.append(f"注意(活動の流れを処理中): {minute_str}は整数ではありません。 処理をスキップします。")
      continue
    
    description: str = match.group(3)
    day_flow_list.append(StructDayFlow(hour, minute, description))
  
  return day_flow_list, warnings

    

def splitHomeContacts(raw: str) -> list:
  splitted = raw.split(", ")
  for i in range(len(splitted)):
    splitted[i] = splitted[i].strip()
  return splitted

def splitNearMisses(raw: str) -> list:
  splitted = raw.split(", ")
  for i in range(len(splitted)):
    splitted[i] = splitted[i].strip()
  return splitted

class StructChildDialyProfile:
    def __init__(self, name: str, attendance: Optional[int], abs_reason: Optional[str],
                 medicine: Optional[int], excretion: Optional[int], eating: Optional[int],
                 sleeping: Optional[int], overview: Optional[str]) -> None:
        self.name = name
        self.attendance = attendance
        self.abs_reason = abs_reason
        self.medicine = medicine
        self.excretion = excretion
        self.eating = eating
        self.sleeping = sleeping
        self.overview = overview
      
    def __repr__(self) -> str:
      return f"StructChildDialyProfile({self.name}, {self.attendance}, {self.abs_reason}, {self.medicine}, {self.excretion}, {self.eating}, {self.sleeping}, {self.overview})"

class ReiwaDate:
  def __init__(self, reiwa: int, month: int, day: int) -> None:
    self.reiwa = reiwa
    self.month = month
    self.day = day
  
  def __hash__(self):
    return hash((self.reiwa, self.month, self.day))

  def __eq__(self, other):
    if isinstance(other, ReiwaDate):
      return self.reiwa == other.reiwa and self.month == other.month and self.day == other.day
    else:
      return False

class StructDiaryPage:
  def __init__(self, dt: ReiwaDate, weather: Optional[str], temperature: Optional[int],
               humidity: Optional[int], recorder: Optional[str], inspections: List[StructInspection], 
               activities: List[str], day_flows: List[StructDayFlow], home_contacts: List[str],
               near_misses: List[str], profiles: Dict[str, StructChildDialyProfile]) -> None:
    self.dt = dt
    self.weather = weather
    self.temperature = temperature
    self.humidity = humidity
    self.recorder = recorder
    self.inspections = inspections
    self.activities = activities
    self.day_flows = day_flows
    self.home_contacts = home_contacts
    self.near_misses = near_misses
    self.profiles = profiles
  
  def __repr__(self) -> str:
    return f"StructDiaryPage({self.weather}, {self.temperature}, {self.humidity}, {self.recorder}, {self.inspections}, {self.activities}, {self.day_flows}, {self.home_contacts}, {self.near_misses}, {self.profiles})"

VALID_ABSENCE_REASONS = {"", "病院受診", "私用", "発熱", "ハナ水・せき", "腹痛", "下痢・嘔吐", "検診", "予防接種", 
      "インフルエンザ", "インフルエンザA", "インフルエンザB", "コロナ感染症", "風しん",
      "水ぼうそう", "おたふく", "プール熱", "結膜炎", "溶蓮菌", "手足口病", "りんご病",
      "RSウイルス", "百日咳", "マイコプラズマ肺炎", "ノロウイルス", "ロタウイルス",
      "ヘルパンギーナ", "突発性発しん", "帯状疱しん", "はしか(麻しん)", "水イボ", "しらみ",
      "とびひ", "その他"}

VALID_WEATHERS = {"晴", "曇", "雨"}
VALID_ATTENDANCE: Dict[str, Optional[int]] = {"---": None, "×": 0, "〇": 2}
VALID_MEDICINE: Dict[str, Optional[int]] = {"---": None, "なし": 0, "こな": 1, "シロップ": 2, "こなとシロップ": 3}
VALID_EXCRETION: Dict[str, Optional[int]] = {"---": None, "×": 0, "△": 1, "〇": 2}
VALID_EATING: Dict[str, Optional[int]] = {"---": None, "×": 0, "△": 1, "〇": 2}
VALID_SLEEPING: Dict[str, Optional[int]] = {"---": None, "×": 0, "△": 1, "〇": 2}

REIWA_NENDO_RANGE = (1, 10)

SHEET_NAME = "Sheet1"
REIWA_POS = (3, 9)
MONTH_POS = (4, 2)
DAY_POS = (4, 4)
WEATHER_POS = (4, 9)
TEMPERATURE_POS = (4, 11)
HUMIDITY_POS = (4, 13)
RECORDER_POS = (5, 10)
INSPECTIONS_POS = (6, 3)
ACTIVITIES_POS = (7, 3)
DAY_FLOWS_POS = (8, 3)
HOME_CONTACTS_POS = (9, 3)
NEAR_MISSES_POS = (10, 3)
PROFILES_POS = (13, 2)

NAME_COL = 2
ATTEND_COL = NAME_COL + 1
REASON_COL = NAME_COL + 2
MEDICINE_COL = NAME_COL + 3
EXCRETION_COL = NAME_COL + 4
EATING_COL = NAME_COL + 6
SLEEPING_COL = NAME_COL + 7
OVERVIEW_COL = NAME_COL + 8

def createDiaryFromXlsxSheet(sheet: Worksheet) -> Tuple[StructDiaryPage, List[str]]:
  warnings: List[str] = []
  raw_reiwa = sheet.cell(REIWA_POS[0], REIWA_POS[1]).internal_value
  raw_reiwa = raw_reiwa if raw_reiwa is not None else ""
  sreiwa: str = str(raw_reiwa)
  try:
    reiwa: int = int(sreiwa)
  except ValueError as e:
    raise AppError(f"エラー(令和年度を解析中): `{sreiwa}`は整数ではありません。")
  if not REIWA_NENDO_RANGE[0] <= reiwa <= REIWA_NENDO_RANGE[1]:
    raise AppError(f"エラー(令和年度を解析中): `{reiwa}`は{REIWA_NENDO_RANGE[0]}以上{REIWA_NENDO_RANGE[1]}以下の整数である必要があります。")
  raw_month = sheet.cell(MONTH_POS[0], MONTH_POS[1]).value
  raw_month = raw_month if raw_month is not None else ""
  smonth = str(raw_month)
  try:
    month: int = int(smonth)
  except ValueError:
    raise AppError(f"エラー(月を解析中): `{smonth}`は整数ではありません。")
  if not 1 <= month <= 12:
    raise AppError(f"エラー(月を解析中): `{month}`は1以上12以下の整数である必要があります。")
  raw_day = sheet.cell(DAY_POS[0], DAY_POS[1]).value
  raw_day = raw_day if raw_day is not None else ""
  sday = str(raw_day)
  try:
    day: int = int(sday)
  except ValueError:
    raise AppError(f"エラー(日を解析中): `{sday}`は整数ではありません。")
  max_day = getNumberOfDays(reiwa, month)
  if not 1 <= day <= max_day:
    raise AppError(f"エラー(日を解析中): `{day}`は1以上{max_day}以下の整数である必要があります。")
  raw_weather = sheet.cell(WEATHER_POS[0], WEATHER_POS[1]).value
  raw_weather = raw_weather if raw_weather is not None else ""
  weather = str(raw_weather).strip()
  if weather.isspace() or weather == "":
    weather = None
  elif weather not in VALID_WEATHERS:
    warnings.append(f"注意(天気を解析中): `{weather}`は有効な天気ではありません。未選択として処理します。")
    weather = None
  raw_temperature = sheet.cell(TEMPERATURE_POS[0], TEMPERATURE_POS[1]).value
  raw_temperature = raw_temperature if raw_temperature is not None else ""
  stemperature = str(raw_temperature).strip()
  stemperature_pattern = r"(\d+)℃"
  stemperature_match = re.match(stemperature_pattern, stemperature)
  if stemperature_match is not None:
    temperature = int(stemperature_match.group(1))
  else:
    warnings.append(f"注意(気温を解析中): `{stemperature}`は`[数字]℃`のフォーマットではなく無効です。0℃として処理します。")
    temperature = 0
  if not (0 <= temperature <= 39):
    warnings.append(f"注意(気温を解析中): `{temperature}`は有効な気温ではありません。0として処理します。")
    temperature = 0
  raw_humidity = sheet.cell(HUMIDITY_POS[0], HUMIDITY_POS[1]).value
  raw_humidity = raw_humidity if raw_humidity is not None else ""
  shumidity = str(raw_humidity).strip()
  shumidity_pattern = r"(\d+)％"
  shumidity_match = re.match(shumidity_pattern, shumidity)
  if shumidity_match is not None:
    humidity = int(shumidity_match.group(1))
  else:
    warnings.append(f"注意(湿度を解析中): `{shumidity}`は`[数字]％`のフォーマットではなく無効です。20％として処理します。")
    humidity = 20
  if not (20 <= humidity <= 100):
    warnings.append(f"注意(湿度を解析中): `{humidity}`は有効な湿度ではありません。20として処理します。")
    humidity = 20
  raw_recorder = sheet.cell(RECORDER_POS[0], RECORDER_POS[1]).value
  raw_recorder = raw_recorder if raw_recorder is not None else ""
  recorder = str(raw_recorder)
  if recorder.isspace() or raw_recorder == "":
    recorder = None
  rawinspections = sheet.cell(INSPECTIONS_POS[0], INSPECTIONS_POS[1]).value
  rawinspections = rawinspections if rawinspections is not None else ""
  sinspections = str(rawinspections)
  inspections = []
  if sinspections != "" and not sinspections.isspace():
    inspections, inspections_warnings = splitInspections(sinspections)
    if len(inspections_warnings) > 0:
      warnings.append(f"視診を解析中に以下の警告がありました：\n" + "\n".join(inspections_warnings))
  rawactivities = sheet.cell(ACTIVITIES_POS[0], ACTIVITIES_POS[1]).value
  rawactivities = rawactivities if rawactivities is not None else ""
  sactivities = str(rawactivities)
  activities = []
  if sactivities != "" and not sactivities.isspace():
    activities = splitActivities(sactivities)
  rawdayflows = sheet.cell(DAY_FLOWS_POS[0], DAY_FLOWS_POS[1]).value
  rawdayflows = rawdayflows if rawdayflows is not None else ""
  sdayflows = str(rawdayflows)
  day_flows = []
  if sdayflows != "" and not sdayflows.isspace():
    day_flows, day_flows_warnings = splitDayFlows(sdayflows)
    if len(day_flows_warnings) > 0:
      warnings.append(f"活動の流れを解析中に以下の警告がありました：\n" + "\n".join(day_flows_warnings))
  rawhomecontacts = sheet.cell(HOME_CONTACTS_POS[0], HOME_CONTACTS_POS[1]).value
  rawhomecontacts = rawhomecontacts if rawhomecontacts is not None else ""
  shomecontacts = str(rawhomecontacts)
  home_contacts = []
  if shomecontacts != "" and not shomecontacts.isspace():
    home_contacts = splitHomeContacts(shomecontacts)
  rawnearmisses = sheet.cell(NEAR_MISSES_POS[0], NEAR_MISSES_POS[1]).value
  rawnearmisses = rawnearmisses if rawnearmisses is not None else ""
  snearmisses = str(rawnearmisses)
  near_misses = []
  if snearmisses is not None and snearmisses != "" and not snearmisses.isspace():
    near_misses = splitNearMisses(snearmisses)
  
  profiles: Dict[str, StructChildDialyProfile] = {}
  current_child_index = 0
  while True:
    raw_name = sheet.cell(PROFILES_POS[0] + current_child_index, NAME_COL).value
    raw_name = raw_name if raw_name is not None else ""
    name = str(raw_name)
    if name is None or name.isspace() or name == "":
      break
    raw_attendance = sheet.cell(PROFILES_POS[0] + current_child_index, ATTEND_COL).value
    raw_attendance = raw_attendance if raw_attendance is not None else ""
    attendance = str(raw_attendance).strip()
    if attendance not in VALID_ATTENDANCE:
      warnings.append(f"注意(出欠を解析中): `{attendance}`は有効な出欠ではありません。未選択として処理します。")
      attendance = "---"
    raw_abs_reason = sheet.cell(PROFILES_POS[0] + current_child_index, REASON_COL).value
    raw_abs_reason = raw_abs_reason if raw_abs_reason is not None else ""
    abs_reason = str(raw_abs_reason).strip()
    if abs_reason == "":
      abs_reason = None
    elif abs_reason not in VALID_ABSENCE_REASONS:
      warnings.append(f"注意(欠席理由を解析中): `{abs_reason}`は有効な欠席理由ではありません。未選択として処理します。")
      abs_reason = None
    raw_medicine = sheet.cell(PROFILES_POS[0] + current_child_index, MEDICINE_COL).value
    raw_medicine = raw_medicine if raw_medicine is not None else ""
    medicine = str(raw_medicine).strip()
    if medicine not in VALID_MEDICINE:
      warnings.append(f"注意(くすりを解析中): `{medicine}`は有効な薬ではありません。未選択として処理します。")
      medicine = "---"
    raw_excretion = sheet.cell(PROFILES_POS[0] + current_child_index, EXCRETION_COL).value
    raw_excretion = raw_excretion if raw_excretion is not None else ""
    excretion = str(raw_excretion).strip()
    if excretion not in VALID_EXCRETION:
      warnings.append(f"注意(排泄を解析中): `{excretion}`は有効な排泄ではありません。未選択として処理します。")
      excretion = "---"
    raw_eating = sheet.cell(PROFILES_POS[0] + current_child_index, EATING_COL).value
    raw_eating = raw_eating if raw_eating is not None else ""
    eating = str(raw_eating).strip()
    if eating not in VALID_EATING:
      warnings.append(f"注意(食事を解析中): `{eating}`は有効な食事ではありません。未選択として処理します。")
      eating = "---"
    raw_sleeping = sheet.cell(PROFILES_POS[0] + current_child_index, SLEEPING_COL).value
    raw_sleeping = raw_sleeping if raw_sleeping is not None else ""
    sleeping = str(raw_sleeping).strip()
    if sleeping not in VALID_SLEEPING:
      warnings.append(f"注意(睡眠を解析中): `{sleeping}`は有効な睡眠ではありません。未選択として処理します。")
      sleeping = "---"
    raw_overview = sheet.cell(PROFILES_POS[0] + current_child_index, OVERVIEW_COL).value
    raw_overview = raw_overview if raw_overview is not None else ""
    overview = str(raw_overview).strip()
    if overview == "" or overview.isspace():
      overview = None
    profile = StructChildDialyProfile(
      name, VALID_ATTENDANCE[attendance], abs_reason, VALID_MEDICINE[medicine], VALID_EXCRETION[excretion], VALID_EATING[eating], VALID_SLEEPING[sleeping], overview
    )
    if name in profiles:
      warnings.append(f"注意(園児プロフィールを解析中): `{name}`は重複しています。このプロフィールをスキップします。")
    profiles[name] = profile
    current_child_index += 1
  dt = ReiwaDate(reiwa, month, day)
  return StructDiaryPage(dt, weather, temperature, humidity, recorder, inspections, activities, day_flows, home_contacts, near_misses, profiles), warnings

def parseXlsx(xlsx_path: Path) -> Tuple[StructDiaryPage, List[str]]:
  if not xlsx_path.exists():
    raise FileNotFoundError(f"ファイルが見つかりません。: {xlsx_path}")
  wb: Workbook = openpyxl.load_workbook(str(xlsx_path))
  sheet: Worksheet = wb[SHEET_NAME]
  try:
    page, warnings = createDiaryFromXlsxSheet(sheet)
  except AppError:
    raise
  finally:
    wb.close()
    
  return page, warnings

def listenPage(page: StructDiaryPage) -> List[Any]:
    n_belongs = len(page.profiles)
    n_attends = 0
    n_absents = 0
    # count attends
    for child_profile in page.profiles.values():
      if child_profile.attendance == 2:
        n_attends += 1
      elif child_profile.attendance == 0:
        n_absents += 1
    
    profile_dumps: List[Any] = []
    for profile in page.profiles.values():
        profile_dump = [
            profile.name, 
            profile.attendance, 
            profile.abs_reason, 
            profile.medicine,
            profile.excretion, 
            profile.eating, 
            profile.sleeping, 
            profile.overview
                ]
        profile_dumps.append(profile_dump)
    dump: List[Any] = [
        page.weather, 
        page.temperature, 
        page.humidity, 
        [n_belongs, n_attends, n_absents],
        page.recorder,
        [[inspection.name, inspection.description] for inspection in page.inspections],
        page.activities,
        [[day_flow.hour, day_flow.minute, day_flow.description] for day_flow in page.day_flows],
        page.home_contacts,
        page.near_misses,
        profile_dumps
    ]
    return dump

def loadListenDiaryPages(xlsx_paths: List[Path]) -> Dict[int, Dict[int, Dict[int, List[Any]]]]:
  listen_diary_pages: Dict[int, Dict[int, Dict[int, List[Any]]]] = {}
  for xlsx_path in xlsx_paths:
    print(f" ----- {xlsx_path.name}を解析中 ----- ")
    diary_page, warnings = parseXlsx(xlsx_path)
    if len(warnings) > 0:
      print("解析を完了しましたが、以下の警告があります。")
      for warning in warnings:
        print(warning)
    if diary_page.dt.reiwa not in listen_diary_pages:
      listen_diary_pages[diary_page.dt.reiwa] = {}
    if diary_page.dt.month not in listen_diary_pages[diary_page.dt.reiwa]:
      listen_diary_pages[diary_page.dt.reiwa][diary_page.dt.month] = {}
    if diary_page.dt.day in listen_diary_pages[diary_page.dt.reiwa][diary_page.dt.month]:
      raise AppError("エラー: 令和{diary_page.dt.reiwa}年{diary_page.dt.month}月{diary_page.dt.day}日の日誌が重複しています。")
    listen_diary_pages[diary_page.dt.reiwa][diary_page.dt.month][diary_page.dt.day] = listenPage(diary_page)
  return listen_diary_pages

def createSaveData(reiwa: int, month: int, day_savedata_map: Dict[int, List[Any]]) -> List[Optional[List[Any]]]:
  save_data: List[Optional[List[Any]]] = [None] * getNumberOfDays(reiwa, month)
  for day in day_savedata_map:
    save_data[day - 1] = day_savedata_map[day]
  return save_data

def margeSaveData(a: List[Optional[List[Any]]], b: List[Optional[List[Any]]]) -> List[Optional[List[Any]]]:
  if len(a) != len(b):
    raise ValueError("リストの長さが異なります。")
  marged: List[Optional[List[Any]]] = [None] * len(a)
  for i in range(len(marged)):
    if a[i] is not None and b[i] is not None:
      raise AppError(f"エラー: {i+1}日のデータが重複しています。")
    elif a[i] is not None:
      marged[i] = a[i]
    elif b[i] is not None:
      marged[i] = b[i]
  return marged
  
def main():
  checkPathsExistence()
  klass: str = loadKlassName(KLASS_NAME_FILE_PATH)
  
  xlsx_path_list: List[Path] = []
  # list existing xlsx files in DIR_PATH
  for path in DIR_PATH.iterdir():
    if path.suffix == ".xlsx":
      xlsx_path_list.append(path.resolve())
  print("=== === === === === エクセルファイル読み込み中 === === === === ===\n")
  listen_diary_pages = loadListenDiaryPages(xlsx_path_list)
  
  save_data_map: Dict[int, Dict[int, List[Optional[List[Any]]]]] = {}
  
  print(f"\n=== === === === === 新しいデータの作成中 === === === === ===\n")
  for reiwa in listen_diary_pages:
    for month in listen_diary_pages[reiwa]:
      print(f" ----- 令和{reiwa}年{month}月のデータを作成中 ----- ")
      xlsx_parsed_data = createSaveData(reiwa, month, listen_diary_pages[reiwa][month])
      existing_data_path = STORAGE_DIR_PATH.joinpath(str(reiwa), klass, f"{month}.json")
      existing_data: Optional[List[Optional[List[Any]]]] = None
      if existing_data_path.exists() and os.path.getsize(existing_data_path) > 0:
        with open(str(existing_data_path), "r", encoding="utf-8") as f:
          existing_data = json.load(f)
      if reiwa not in save_data_map:
        save_data_map[reiwa] = {}
      if existing_data is not None:
        marged = margeSaveData(existing_data, xlsx_parsed_data)
        save_data_map[reiwa][month] = marged
      else:
        save_data_map[reiwa][month] = xlsx_parsed_data
  
  print(f"\n=== === === === === データの保存中 === === === === ===\n")
  for reiwa in save_data_map:
    for month in save_data_map[reiwa]:
      print(f" ----- 令和{reiwa}年{month}月のデータを保存中 ----- ")
      dumped = json.dumps(save_data_map[reiwa][month], ensure_ascii=False, indent=5)
      if not STORAGE_DIR_PATH.joinpath(str(reiwa)).exists():
        STORAGE_DIR_PATH.joinpath(str(reiwa)).mkdir()
      if not STORAGE_DIR_PATH.joinpath(str(reiwa), klass).exists():
        STORAGE_DIR_PATH.joinpath(str(reiwa), klass).mkdir()
        for i in range(1, 13):
          STORAGE_DIR_PATH.joinpath(str(reiwa), klass, f"{i}.json").touch()
      file_path = STORAGE_DIR_PATH.joinpath(str(reiwa), klass, f"{month}.json")
      if not file_path.exists():
        raise FileNotFoundError(f"令和{reiwa}年{month}月のデータが存在しません。")
      with open(str(file_path), "w", encoding="utf-8") as f:
        f.write(dumped)

  print("\n\n完了しました。\n")

if __name__ == "__main__":
  try:
    main()
  except AppError as e:
    print(e.message)
    abortProgram()